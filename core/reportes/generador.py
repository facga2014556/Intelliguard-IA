from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
import os

class GeneradorReportes:
    def __init__(self, db):
        self.db = db
        
    def generar_reporte_pertenencias(self, filtros=None):
        """
        Genera un reporte Excel de pertenencias
        
        Args:
            filtros: Diccionario con filtros opcionales (estudiante, estado, etc)
            
        Returns:
            BytesIO: Buffer con el archivo Excel
        """
        try:
            # Obtener datos de la base de datos
            query = """
                SELECT 
                    rp.idPertenencia,
                    e.codigoEstudiante,
                    e.Nombres,
                    o.Nombre as nombreObjeto,
                    rp.Hora_Entrada,
                    rp.Hora_Salida,
                    ep.estado
                FROM registros_pertenencia rp
                JOIN estudiantes e ON rp.idEstudiante = e.idEstudiante
                JOIN objetos o ON rp.idObjeto = o.idObjeto
                JOIN estado_pertenencias ep ON rp.idEstado = ep.id
                WHERE 1=1
            """
            
            # Aplicar filtros si existen
            params = []
            if filtros:
                if filtros.get('estudiante'):
                    query += " AND (e.Nombres LIKE %s OR e.codigoEstudiante LIKE %s)"
                    params.extend([f"%{filtros['estudiante']}%", f"%{filtros['estudiante']}%"])
                if filtros.get('estado'):
                    query += " AND ep.estado LIKE %s"
                    params.append(f"%{filtros['estado']}%")
                    
            # Ejecutar consulta
            resultados = self.db.obtener_todos(query, tuple(params))
            
            # Crear Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Pertenencias"
            
            # Estilo para encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Agregar encabezados
            headers = ['ID', 'Código Estudiante', 'Nombre Estudiante', 'Objeto', 
                      'Hora Entrada', 'Hora Salida', 'Estado']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Agregar datos
            for row, r in enumerate(resultados, 2):
                ws.append([
                    r[0],  # ID
                    r[1],  # Código Estudiante
                    r[2],  # Nombre Estudiante
                    r[3],  # Objeto
                    r[4],  # Hora Entrada
                    r[5],  # Hora Salida
                    r[6]   # Estado
                ])
                
                # Centrar todas las celdas
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            
            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Guardar en buffer
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            print(f"Error al generar reporte: {str(e)}")
            raise 